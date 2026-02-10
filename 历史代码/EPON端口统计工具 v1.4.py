import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
import queue
import os
import re
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import traceback
from datetime import datetime

class EPONPortAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("EPON端口统计工具 v1.4")
        self.root.geometry("900x750")
        self.root.minsize(800, 600)
        
        self.log_queue = queue.Queue()
        self.create_widgets()
        self.update_log()
    
    def create_widgets(self):
        main_frame = tk.Frame(self.root, padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 输入区域
        input_frame = tk.LabelFrame(main_frame, text="输入设置", padx=10, pady=10)
        input_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(input_frame, text="输入路径:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.input_path_var = tk.StringVar()
        tk.Entry(input_frame, textvariable=self.input_path_var, width=60).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        tk.Button(input_frame, text="📁 选择文件", command=self.select_input_file, width=12).grid(row=0, column=2, padx=5, pady=5)
        tk.Button(input_frame, text="📁 选择文件夹", command=self.select_input_folder, width=12).grid(row=0, column=3, padx=5, pady=5)
        input_frame.columnconfigure(1, weight=1)
        
        # 输出区域
        output_frame = tk.LabelFrame(main_frame, text="输出设置", padx=10, pady=10)
        output_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(output_frame, text="输出目录:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.output_dir_var = tk.StringVar()
        tk.Entry(output_frame, textvariable=self.output_dir_var, width=60).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        tk.Button(output_frame, text="📁 选择目录", command=self.select_output_dir, width=12).grid(row=0, column=2, padx=5, pady=5)
        output_frame.columnconfigure(1, weight=1)
        
        # 日志区域
        log_frame = tk.LabelFrame(main_frame, text="处理日志", padx=10, pady=10)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, wrap=tk.WORD, font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # 按钮区域
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 5))
        self.process_btn = tk.Button(btn_frame, text="🚀 开始处理并自动汇总", command=self.start_processing, bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), height=2)
        self.process_btn.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        tk.Button(btn_frame, text="🧹 清空日志", command=self.clear_log, width=12).pack(side=tk.RIGHT, padx=5)
        
        self.status_var = tk.StringVar(value="就绪")
        tk.Label(self.root, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W, padx=10).pack(side=tk.BOTTOM, fill=tk.X)

    def select_input_file(self):
        file_path = filedialog.askopenfilename(title="选择EPON数据文件", filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")])
        if file_path: self.input_path_var.set(file_path)

    def select_input_folder(self):
        folder_path = filedialog.askdirectory(title="选择包含EPON数据文件的文件夹")
        if folder_path: self.input_path_var.set(folder_path)

    def select_output_dir(self):
        dir_path = filedialog.askdirectory(title="选择输出目录")
        if dir_path: self.output_dir_var.set(dir_path)

    def clear_log(self):
        self.log_text.delete(1.0, tk.END)

    def log(self, message):
        self.log_queue.put(message)

    def update_log(self):
        try:
            while True:
                message = self.log_queue.get_nowait()
                self.log_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
                self.log_text.see(tk.END)
        except queue.Empty: pass
        self.root.after(100, self.update_log)

    def start_processing(self):
        input_path = self.input_path_var.get().strip()
        output_dir = self.output_dir_var.get().strip()
        if not input_path or not os.path.exists(input_path):
            messagebox.showerror("错误", "输入路径无效！")
            return
        if not output_dir or not os.path.exists(output_dir):
            messagebox.showerror("错误", "输出目录无效！")
            return
        
        self.process_btn.config(state=tk.DISABLED, text="⏳ 正在统计数据...")
        self.status_var.set("正在处理...")
        threading.Thread(target=self.process_task, args=(input_path, output_dir), daemon=True).start()

    def process_task(self, input_path, output_dir):
        try:
            files = [input_path] if os.path.isfile(input_path) else [os.path.join(input_path, f) for f in os.listdir(input_path) if f.lower().endswith('.txt')]
            if not files: raise ValueError("未找到TXT文件")

            for file_path in files:
                self.log(f"解析文件: {os.path.basename(file_path)}")
                slot_data = self.parse_epon_data(file_path)
                output_filename = os.path.splitext(os.path.basename(file_path))[0] + ".xlsx"
                self.generate_excel_report(slot_data, os.path.join(output_dir, output_filename))
            
            self.log(f"处理完成！")
            self.root.after(0, lambda: messagebox.showinfo("完成", "报表生成完毕。"))
            self.root.after(0, lambda: os.startfile(output_dir))
        except Exception as e:
            self.log(f"严重错误: {traceback.format_exc()}")
        finally:
            self.root.after(0, lambda: self.process_btn.config(state=tk.NORMAL, text="🚀 开始处理"))
            self.root.after(0, lambda: self.status_var.set("就绪"))

    def parse_epon_data(self, file_path):
        slot_data = {s: {p: {'在线': 0, '离线': 0, '静默': 0} for p in range(1, 25)} for s in range(2, 8)}
        current_slot, current_pon = None, None
        content = None
        for enc in ['utf-8', 'gbk', 'gb2312']:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    content = f.readlines()
                break
            except: continue
        if not content: raise ValueError("无法读取文件")

        for line in content:
            line = line.strip()
            if 'dis onu slot' in line:
                match = re.search(r'dis onu slot\s+(\d+)', line)
                if match: current_slot = int(match.group(1))
                continue
            if current_slot and 2 <= current_slot <= 7 and 'Olt' in line and '/0/' in line:
                match = re.search(r'Olt\d+/0/(\d+)', line)
                if match: current_pon = int(match.group(1))
                continue
            if current_slot and current_pon and line and not line.startswith('-'):
                if any(k in line for k in ['State', 'MAC', 'LOID', 'LLID', 'Port']): continue
                parts = re.split(r'\s+', line)
                if len(parts) >= 2:
                    state = parts[-2]
                    key = '在线' if state == 'Up' else '离线' if state == 'Offline' else '静默' if state == 'Silent' else None
                    if key: slot_data[current_slot][current_pon][key] += 1
        return slot_data

    def generate_excel_report(self, slot_data, output_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EPON统计报表"

        # 样式定义
        color_slot_name = PatternFill(start_color="FDE9D9", fill_type="solid")
        color_pon_header = PatternFill(start_color="D9E1F2", fill_type="solid")
        color_idle_yes = PatternFill(start_color="FFFF00", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal="center", vertical="center")

        total_idle_count = 0

        # 1. 标题
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f"统计信息(生成日期: {datetime.now().strftime('%Y-%m-%d')})"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = center_align

        # 2. 列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        for col in range(3, 15):
            ws.column_dimensions[get_column_letter(col)].width = 5

        current_row = 2
        slot_names = {2: "2号槽位", 3: "3号槽位", 4: "4号槽位", 5: "5号槽位", 6: "6号槽位", 7: "7号槽位"}

        # 3. 槽位数据循环
        for slot_num in range(2, 8):
            start_merge_row = current_row
            slot_info = slot_data.get(slot_num, {})
            
            for group in [range(1, 24, 2), range(2, 25, 2)]:
                rows = [("PON", None), ("在线", "在线"), ("离线", "离线"), ("静默", "静默"), ("空闲", "空闲")]
                for label, data_key in rows:
                    ws.cell(row=current_row, column=2, value=label).alignment = center_align
                    for idx, pon_id in enumerate(group, start=3):
                        cell = ws.cell(row=current_row, column=idx)
                        cell.alignment = center_align
                        if label == "PON":
                            cell.value = pon_id
                            cell.fill = color_pon_header
                            ws.cell(row=current_row, column=2).fill = color_pon_header
                        elif label == "空闲":
                            is_idle = slot_info.get(pon_id, {}).get('在线', 0) == 0
                            if is_idle:
                                cell.value = "是"
                                cell.fill = color_idle_yes
                                cell.font = Font(bold=True)
                                total_idle_count += 1
                            else:
                                cell.value = "否"
                        else:
                            cell.value = slot_info.get(pon_id, {}).get(data_key, 0)
                    current_row += 1
            
            ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=current_row-1, end_column=1)
            slot_cell = ws.cell(row=start_merge_row, column=1, value=slot_names[slot_num])
            slot_cell.alignment = center_align
            slot_cell.fill = color_slot_name
            slot_cell.font = Font(bold=True)

        # 4. 给表格主体添加边框 (不包含底部的统计行)
        for r in range(1, current_row):
            for c in range(1, 15):
                ws.cell(row=r, column=c).border = thin_border

        # 5. 修改后的统计行：移除 stat_cell.border 赋值
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        stat_cell = ws.cell(row=current_row, column=1)
        stat_cell.value = f"截止2026年02月10日统计该设备已回收PON口数量：{total_idle_count}"
        stat_cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")
        stat_cell.alignment = center_align
        stat_cell.font = Font(bold=True)
        # 注意：此处不再设置 border，使其没有黑框线

        # 6. 备注
        notes = [
            "", "备注：",
            "1. 空闲一栏标记为“是”，说明该口下无在线用户。需留意离线和静默数量。",
            "2. 离线：若确认为撤销点位请反馈技术部删除配置；FTTH日常关机则无需处理。",
            "3. 静默：说明有ONU在线但未配置业务，请及时核实并下发配置。",
            "4. 统计结果以发布日期当天为准。"
        ]
        for note in notes:
            current_row += 1
            ws.cell(row=current_row, column=1, value=note).font = Font(size=10)

        wb.save(output_path)

if __name__ == "__main__":
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except: pass
    root = tk.Tk()
    app = EPONPortAnalyzer(root)
    root.mainloop()