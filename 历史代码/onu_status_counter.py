#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
ONU状态统计脚本
用于统计7606-10.txt文件中每个OLT端口ONU的在线、离线和静默数量
并将结果输出到Excel表格中

作者: Freak
版本: 1.0.0
日期: 2025-07-01
"""

import re
import os
import sys
import datetime
import argparse
import logging
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 缺少必要的库 'openpyxl'，请使用以下命令安装:")
    print("pip install openpyxl")
    sys.exit(1)


def parse_onu_data(file_path):
    """
    解析ONU数据文件，统计每个OLT端口的ONU状态
    
    Args:
        file_path: ONU数据文件路径
        
    Returns:
        dict: 包含每个OLT端口ONU状态统计的字典
    """
    # 初始化结果字典，使用defaultdict避免键不存在的问题
    olt_stats = defaultdict(lambda: {'Up': 0, 'Offline': 0, 'Silent': 0})
    
    # 当前正在处理的OLT端口
    current_olt = None
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            
            # 查找静默ONU的信息
            silent_onus = set()
            silent_mode = False
            for i, line in enumerate(lines):
                if 'dis onu silent' in line:
                    silent_mode = True
                    continue
                    
                if silent_mode and '---------------------------------- Olt' in line:
                    silent_mode = False
                    
                if silent_mode and re.match(r'\w{4}-\w{4}-\w{4}\s+Onu\d+/\d+/\d+:\d+', line):
                    # 提取静默ONU的端口信息
                    match = re.search(r'(Onu\d+/\d+/\d+):\d+', line)
                    if match:
                        port = match.group(1)
                        silent_onus.add(port)
            
            # 解析OLT端口和ONU状态信息
            for line in lines:
                # 匹配OLT端口行
                olt_match = re.search(r'-- (Olt\d+/\d+/\d+) --', line)
                if olt_match:
                    current_olt = olt_match.group(1)
                    continue
                
                # 如果当前有OLT端口，且行包含ONU信息，则解析状态
                if current_olt and re.match(r'\w{4}-\w{4}-\w{4}\s+Onu', line):
                    # 提取状态信息
                    state_match = re.search(r'\s+(Up|Offline)\s+', line)
                    if state_match:
                        state = state_match.group(1)
                        olt_stats[current_olt][state] += 1
                
                # 检查是否有总结行，如果有则重置当前OLT
                if line.strip().startswith('ONUs found:'):
                    current_olt = None
            
            # 处理静默ONU
            for port in silent_onus:
                # 从Onu端口提取对应的Olt端口
                olt_port = port.replace('Onu', 'Olt')
                if olt_port in olt_stats:
                    olt_stats[olt_port]['Silent'] += 1
        
        return olt_stats
    
    except Exception as e:
        print(f"解析文件时出错: {e}")
        return {}


def print_statistics(olt_stats):
    """
    打印OLT端口ONU状态统计
    
    Args:
        olt_stats: 包含每个OLT端口ONU状态统计的字典
    """
    print("\n统计结果:")
    print("-" * 50)
    print(f"{'OLT端口':<15}{'在线':<10}{'离线':<10}{'静默':<10}{'总数':<10}")
    print("-" * 50)
    
    total_up = 0
    total_offline = 0
    total_silent = 0
    
    # 按OLT端口排序
    for olt in sorted(olt_stats.keys()):
        stats = olt_stats[olt]
        up = stats['Up']
        offline = stats['Offline']
        silent = stats['Silent']
        total = up + offline + silent
        
        print(f"{olt:<15}{up:<10}{offline:<10}{silent:<10}{total:<10}")
        
        total_up += up
        total_offline += offline
        total_silent += silent
    
    total_all = total_up + total_offline + total_silent
    print("-" * 50)
    print(f"{'总计':<15}{total_up:<10}{total_offline:<10}{total_silent:<10}{total_all:<10}")


def create_excel_report(olt_stats, output_file):
    """
    创建Excel报表
    
    Args:
        olt_stats: 包含每个OLT端口ONU状态统计的字典
        output_file: 输出Excel文件路径
    """
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ONU状态统计"
    
    # 设置单元格样式
    title_font = Font(name='宋体', size=14, bold=True)
    header_font = Font(name='宋体', size=11, bold=True)
    cell_font = Font(name='宋体', size=11)
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 填充颜色
    blue_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    orange_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    
    # 居中对齐
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列宽
    for col in range(1, 25):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # 添加标题
    current_date = datetime.datetime.now().strftime("%Y. %m. %d")
    title = f"日报7606-10 （{current_date}）"
    ws.merge_cells('A1:Y1')  # 扩展到25列
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    
    # 为标题单元格添加边框
    for col in range(1, 26):  # 扩展到25列
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
    
    # 组织数据按槽位分组
    slot_data = {}
    for olt in olt_stats:
        # 提取槽位号，格式为Olt{slot}/0/{port}
        match = re.match(r'Olt(\d+)/\d+/(\d+)', olt)
        if match:
            slot = int(match.group(1))
            port = int(match.group(2))
            
            if slot not in slot_data:
                slot_data[slot] = {}
            
            slot_data[slot][port] = olt_stats[olt]
    
    # 当前行索引
    row = 2
    
    # 为每个槽位创建表格
    for slot in sorted(slot_data.keys()):
        # 添加槽位标题
        ws.merge_cells(f'A{row}:Y{row}')
        ws[f'A{row}'] = f"{slot}号槽位"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_alignment
        
        # 设置背景色
        fill = blue_fill if slot % 2 == 0 else orange_fill
        for col in range(1, 26):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
        
        row += 1
        
        # 添加PON口行
        ws[f'A{row}'] = "PON"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].border = thin_border
        
        # 添加奇数PON口编号
        for i, col in enumerate(range(2, 25, 2)):
            port = i * 2 + 1  # 1, 3, 5, ...
            ws.cell(row=row, column=col).value = port
            ws.cell(row=row, column=col).font = cell_font
            ws.cell(row=row, column=col).alignment = center_alignment
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        
        # 添加在线行
        ws[f'A{row}'] = "在线"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].border = thin_border
        
        # 添加奇数PON口在线数量
        for i, col in enumerate(range(2, 25, 2)):
            port = i * 2 + 1  # 1, 3, 5, ...
            if port in slot_data[slot]:
                ws.cell(row=row, column=col).value = slot_data[slot][port]['Up']
            else:
                ws.cell(row=row, column=col).value = ""
            ws.cell(row=row, column=col).font = cell_font
            ws.cell(row=row, column=col).alignment = center_alignment
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        
        # 添加离线行
        ws[f'A{row}'] = "离线"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].border = thin_border
        
        # 添加奇数PON口离线数量
        for i, col in enumerate(range(2, 25, 2)):
            port = i * 2 + 1  # 1, 3, 5, ...
            if port in slot_data[slot]:
                ws.cell(row=row, column=col).value = slot_data[slot][port]['Offline']
            else:
                ws.cell(row=row, column=col).value = ""
            ws.cell(row=row, column=col).font = cell_font
            ws.cell(row=row, column=col).alignment = center_alignment
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        
        # 添加静默行
        ws[f'A{row}'] = "静默"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].border = thin_border
        
        # 添加奇数PON口静默数量
        for i, col in enumerate(range(2, 25, 2)):
            port = i * 2 + 1  # 1, 3, 5, ...
            if port in slot_data[slot]:
                ws.cell(row=row, column=col).value = slot_data[slot][port]['Silent']
            else:
                ws.cell(row=row, column=col).value = "0"
            ws.cell(row=row, column=col).font = cell_font
            ws.cell(row=row, column=col).alignment = center_alignment
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        
        # 添加空闲行
        ws[f'A{row}'] = "空闲"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].border = thin_border
        
        # 添加奇数PON口空闲状态
        for i, col in enumerate(range(2, 25, 2)):
            port = i * 2 + 1  # 1, 3, 5, ...
            if port in slot_data[slot]:
                total = slot_data[slot][port]['Up'] + slot_data[slot][port]['Offline'] + slot_data[slot][port]['Silent']
                ws.cell(row=row, column=col).value = "否" if total > 0 else "是"
            else:
                ws.cell(row=row, column=col).value = "是"
            ws.cell(row=row, column=col).font = cell_font
            ws.cell(row=row, column=col).alignment = center_alignment
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        
        # 添加PON口行（偶数PON口）
        ws[f'A{row}'] = "PON"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].border = thin_border
        
        # 添加偶数PON口编号
        for i, col in enumerate(range(2, 25, 2)):
            port = (i + 1) * 2  # 2, 4, 6, ...
            ws.cell(row=row, column=col).value = port
            ws.cell(row=row, column=col).font = cell_font
            ws.cell(row=row, column=col).alignment = center_alignment
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        
        # 添加在线行（偶数PON口）
        ws[f'A{row}'] = "在线"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].border = thin_border
        
        # 添加偶数PON口在线数量
        for i, col in enumerate(range(2, 25, 2)):
            port = (i + 1) * 2  # 2, 4, 6, ...
            if port in slot_data[slot]:
                ws.cell(row=row, column=col).value = slot_data[slot][port]['Up']
            else:
                ws.cell(row=row, column=col).value = ""
            ws.cell(row=row, column=col).font = cell_font
            ws.cell(row=row, column=col).alignment = center_alignment
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        
        # 添加离线行（偶数PON口）
        ws[f'A{row}'] = "离线"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].border = thin_border
        
        # 添加偶数PON口离线数量
        for i, col in enumerate(range(2, 25, 2)):
            port = (i + 1) * 2  # 2, 4, 6, ...
            if port in slot_data[slot]:
                ws.cell(row=row, column=col).value = slot_data[slot][port]['Offline']
            else:
                ws.cell(row=row, column=col).value = ""
            ws.cell(row=row, column=col).font = cell_font
            ws.cell(row=row, column=col).alignment = center_alignment
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        
        # 添加静默行（偶数PON口）
        ws[f'A{row}'] = "静默"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].border = thin_border
        
        # 添加偶数PON口静默数量
        for i, col in enumerate(range(2, 25, 2)):
            port = (i + 1) * 2  # 2, 4, 6, ...
            if port in slot_data[slot]:
                ws.cell(row=row, column=col).value = slot_data[slot][port]['Silent']
            else:
                ws.cell(row=row, column=col).value = "0"
            ws.cell(row=row, column=col).font = cell_font
            ws.cell(row=row, column=col).alignment = center_alignment
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        
        # 添加空闲行（偶数PON口）
        ws[f'A{row}'] = "空闲"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].border = thin_border
        
        # 添加偶数PON口空闲状态
        for i, col in enumerate(range(2, 25, 2)):
            port = (i + 1) * 2  # 2, 4, 6, ...
            if port in slot_data[slot]:
                total = slot_data[slot][port]['Up'] + slot_data[slot][port]['Offline'] + slot_data[slot][port]['Silent']
                ws.cell(row=row, column=col).value = "否" if total > 0 else "是"
            else:
                ws.cell(row=row, column=col).value = "是"
            ws.cell(row=row, column=col).font = cell_font
            ws.cell(row=row, column=col).alignment = center_alignment
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    # 保存Excel文件
    wb.save(output_file)
    print(f"Excel报表已保存到: {output_file}")


def setup_logging():
    """
    设置日志记录
    """
    log_format = '%(asctime)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=logging.INFO, format=log_format)


def parse_arguments():
    """
    解析命令行参数
    
    Returns:
        argparse.Namespace: 解析后的命令行参数
    """
    parser = argparse.ArgumentParser(description='ONU状态统计工具')
    parser.add_argument('-i', '--input', help='输入文件路径，默认为脚本目录下的7606-10.txt')
    parser.add_argument('-o', '--output', help='输出Excel文件路径，默认为脚本目录下的ONU状态统计_日期.xlsx')
    parser.add_argument('--no-open', action='store_true', help='不自动打开生成的Excel文件')
    parser.add_argument('--debug', action='store_true', help='启用调试模式')
    
    return parser.parse_args()


def main():
    # 设置日志记录
    setup_logging()
    
    # 解析命令行参数
    args = parse_arguments()
    
    # 设置日志级别
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 构建数据文件路径
    if args.input:
        data_file = args.input
    else:
        data_file = os.path.join(script_dir, '7606-10.txt')
    
    # 检查文件是否存在
    if not os.path.exists(data_file):
        logging.error(f"错误: 文件 '{data_file}' 不存在!")
        return 1
    
    logging.info(f"开始解析文件: {data_file}")
    
    # 解析数据并打印统计结果
    olt_stats = parse_onu_data(data_file)
    if not olt_stats:
        logging.error("解析数据失败!")
        return 1
    
    print_statistics(olt_stats)
    
    # 生成Excel报表
    current_date = datetime.datetime.now().strftime("%Y%m%d")
    if args.output:
        output_file = args.output
    else:
        output_file = os.path.join(script_dir, f"ONU状态统计_{current_date}.xlsx")
    
    logging.info(f"正在生成Excel报表: {output_file}")
    create_excel_report(olt_stats, output_file)
    
    print(f"\n统计完成!")
    print(f"Excel报表已保存到: {output_file}")
    
    # 尝试自动打开Excel文件
    if not args.no_open:
        try:
            os.startfile(output_file)
            logging.info("已自动打开Excel报表")
        except Exception as e:
            logging.warning(f"无法自动打开Excel文件: {e}")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())