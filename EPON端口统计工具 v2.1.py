import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import threading
import queue
import os
import re
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import traceback
from datetime import datetime
import time

# 尝试导入paramiko，如果未安装则给出提示
try:
    import paramiko
    PARAMIKO_AVAILABLE = True
except ImportError:
    PARAMIKO_AVAILABLE = False


class SSHConnection:
    """SSH连接管理类"""
    
    def __init__(self, host, port, username, password):
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        self.client = None
        self.shell = None
        
    def connect(self, timeout=10):
        """建立SSH连接并设置屏幕长度"""
        self.client = paramiko.SSHClient()
        self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        self.client.connect(
            hostname=self.host,
            port=self.port,
            username=self.username,
            password=self.password,
            timeout=timeout,
            look_for_keys=False,
            allow_agent=False
        )
        # 创建交互式shell
        self.shell = self.client.invoke_shell()
        # 等待初始提示符
        time.sleep(1)
        # 清空初始输出
        while self.shell.recv_ready():
            self.shell.recv(65535)
        
        # 设置屏幕长度，确保输出不分页
        self.shell.send('screen-length disable\n')
        time.sleep(0.5)
        # 清空输出
        while self.shell.recv_ready():
            self.shell.recv(65535)
        
        return True
    
    def get_full_output(self, cmd, timeout=30):
        """获取完整命令输出，自动处理分页"""
        if not self.shell:
            raise Exception("SSH未连接")
        
        self.shell.send(cmd + '\n')
        output = ""
        end_time = time.time() + timeout
        last_chunk_time = time.time()
        
        while time.time() < end_time:
            if self.shell.recv_ready():
                chunk = self.shell.recv(65535).decode('utf-8', errors='ignore')
                output += chunk
                last_chunk_time = time.time()
                
                # 检查是否输出完整（出现ONUs found表示完成）
                if "ONUs found:" in chunk:
                    # 再等待一下确保没有更多输出
                    time.sleep(0.5)
                    # 清空剩余输出
                    while self.shell.recv_ready():
                        output += self.shell.recv(65535).decode('utf-8', errors='ignore')
                    return output
                
                # 检查是否需要继续（分页提示）
                if "More" in chunk or "----" in chunk[-50:]:
                    self.shell.send(' ')
                    time.sleep(0.3)
                    continue
                    
            # 如果5秒内没有新数据，认为输出完成
            if time.time() - last_chunk_time > 5:
                break
                
            time.sleep(0.1)
        
        return output
    
    def close(self):
        """关闭SSH连接"""
        if self.shell:
            self.shell.close()
        if self.client:
            self.client.close()


class EPONPortAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("EPON端口统计工具 v2.1 by Freak")
        self.root.geometry("1000x800")
        self.root.minsize(900, 700)
        
        # 配置grid权重使窗口可拉伸
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        self.log_queue = queue.Queue()
        self.stop_flag = False  # 停止标志
        self.ssh_thread = None  # SSH线程引用
        self.create_widgets()
        self.update_log()
        
        # 检查paramiko是否可用
        if not PARAMIKO_AVAILABLE:
            self.log("警告: 未安装paramiko库，SSH功能不可用。请运行: pip install paramiko")
    
    def create_widgets(self):
        # 主框架使用grid布局
        main_frame = tk.Frame(self.root, padx=10, pady=10)
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)  # 日志区域可拉伸
        
        # 创建Notebook（标签页）
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
        
        # ========== SSH采集标签页 ==========
        ssh_frame = tk.Frame(self.notebook, padx=10, pady=10)
        self.notebook.add(ssh_frame, text="SSH远程采集")
        ssh_frame.columnconfigure(0, weight=1)
        ssh_frame.rowconfigure(0, weight=1)
        
        # SSH采集内容框架
        ssh_content = tk.Frame(ssh_frame)
        ssh_content.pack(fill=tk.BOTH, expand=True)
        ssh_content.columnconfigure(1, weight=1)
        
        # 设备列表区域
        device_frame = tk.LabelFrame(ssh_content, text="设备列表 (格式: 设备名-[IP]，每行一个)", padx=10, pady=10)
        device_frame.grid(row=0, column=0, columnspan=3, sticky="nsew", pady=(0, 10))
        device_frame.columnconfigure(0, weight=1)
        device_frame.rowconfigure(0, weight=1)
        
        self.device_text = scrolledtext.ScrolledText(device_frame, height=6, wrap=tk.WORD, font=("Consolas", 10))
        self.device_text.grid(row=0, column=0, sticky="nsew")
        #self.device_text.insert(tk.END, "山下湖-172.10.1.26\n璜山-172.10.1.16\n中心12F-7606-5-双向-172.10.1.1")
        
        # SSH认证区域
        auth_frame = tk.LabelFrame(ssh_content, text="SSH认证信息", padx=10, pady=10)
        auth_frame.grid(row=1, column=0, columnspan=3, sticky="nsew", pady=(0, 10))
        auth_frame.columnconfigure(1, weight=1)
        auth_frame.columnconfigure(3, weight=1)
        
        tk.Label(auth_frame, text="用户名:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.username_var = tk.StringVar()
        tk.Entry(auth_frame, textvariable=self.username_var, width=20).grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        
        tk.Label(auth_frame, text="密码:").grid(row=0, column=2, sticky="w", padx=5, pady=5)
        self.password_var = tk.StringVar()
        tk.Entry(auth_frame, textvariable=self.password_var, width=20, show="*").grid(row=0, column=3, sticky="ew", padx=5, pady=5)
        
        tk.Label(auth_frame, text="端口:").grid(row=0, column=4, sticky="w", padx=5, pady=5)
        self.port_var = tk.StringVar(value="22")
        tk.Entry(auth_frame, textvariable=self.port_var, width=8).grid(row=0, column=5, sticky="ew", padx=5, pady=5)
        
        # 输出目录区域
        output_frame = tk.LabelFrame(ssh_content, text="输出设置", padx=10, pady=10)
        output_frame.grid(row=2, column=0, columnspan=3, sticky="nsew", pady=(0, 10))
        output_frame.columnconfigure(1, weight=1)
        
        tk.Label(output_frame, text="输出目录:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.output_dir_var = tk.StringVar()
        tk.Entry(output_frame, textvariable=self.output_dir_var, width=60).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        tk.Button(output_frame, text="📁 选择目录", command=self.select_output_dir, width=12).grid(row=0, column=2, padx=5, pady=5)
        
        # SSH采集按钮区域
        ssh_btn_frame = tk.Frame(ssh_content)
        ssh_btn_frame.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(0, 5))
        
        self.ssh_btn = tk.Button(ssh_btn_frame, text="🚀 开始SSH采集", command=self.start_ssh_collection, 
                                  bg="#2196F3", fg="white", font=("Arial", 10, "bold"), height=2)
        self.ssh_btn.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        self.stop_ssh_btn = tk.Button(ssh_btn_frame, text="⏹ 停止采集", command=self.stop_ssh_collection, 
                                       bg="#f44336", fg="white", font=("Arial", 10, "bold"), height=2,
                                       state=tk.DISABLED)
        self.stop_ssh_btn.pack(side=tk.RIGHT, padx=5, expand=True, fill=tk.X)
        
        # ========== 本地分析标签页 ==========
        local_frame = tk.Frame(self.notebook, padx=10, pady=10)
        self.notebook.add(local_frame, text="本地文件分析")
        local_frame.columnconfigure(0, weight=1)
        
        # 输入区域
        input_frame = tk.LabelFrame(local_frame, text="输入设置", padx=10, pady=10)
        input_frame.pack(fill=tk.X, pady=(0, 10))
        input_frame.columnconfigure(1, weight=1)
        
        tk.Label(input_frame, text="输入路径:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.input_path_var = tk.StringVar()
        tk.Entry(input_frame, textvariable=self.input_path_var, width=60).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        tk.Button(input_frame, text="📁 选择文件", command=self.select_input_file, width=12).grid(row=0, column=2, padx=5, pady=5)
        tk.Button(input_frame, text="📁 选择文件夹", command=self.select_input_folder, width=12).grid(row=0, column=3, padx=5, pady=5)
        
        # 本地分析输出目录
        local_output_frame = tk.LabelFrame(local_frame, text="输出设置", padx=10, pady=10)
        local_output_frame.pack(fill=tk.X, pady=(0, 10))
        local_output_frame.columnconfigure(1, weight=1)
        
        tk.Label(local_output_frame, text="输出目录:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.local_output_dir_var = tk.StringVar()
        tk.Entry(local_output_frame, textvariable=self.local_output_dir_var, width=60).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        tk.Button(local_output_frame, text="📁 选择目录", command=self.select_local_output_dir, width=12).grid(row=0, column=2, padx=5, pady=5)
        
        # 本地分析按钮
        local_btn_frame = tk.Frame(local_frame)
        local_btn_frame.pack(fill=tk.X, pady=(0, 5))
        
        self.process_btn = tk.Button(local_btn_frame, text="🚀 开始处理并生成Excel", command=self.start_processing, 
                                      bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), height=2)
        self.process_btn.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        # ========== 公共日志区域 ==========
        log_frame = tk.LabelFrame(main_frame, text="处理日志", padx=10, pady=10)
        log_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, font=("Consolas", 9))
        self.log_text.grid(row=0, column=0, sticky="nsew")
        
        # 底部按钮区域
        btn_frame = tk.Frame(main_frame)
        btn_frame.grid(row=2, column=0, sticky="ew")
        
        tk.Button(btn_frame, text="🧹 清空日志", command=self.clear_log, width=15).pack(side=tk.RIGHT, padx=5)
        
        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        tk.Label(self.root, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W, padx=10).grid(row=1, column=0, sticky="ew")

    def select_input_file(self):
        file_path = filedialog.askopenfilename(title="选择EPON数据文件", filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")])
        if file_path: 
            self.input_path_var.set(file_path)
            # 自动设置输出目录为文件所在目录
            if not self.local_output_dir_var.get():
                self.local_output_dir_var.set(os.path.dirname(file_path))

    def select_input_folder(self):
        folder_path = filedialog.askdirectory(title="选择包含EPON数据文件的文件夹")
        if file_path: 
            self.input_path_var.set(file_path)

    def select_input_folder(self):
        folder_path = filedialog.askdirectory(title="选择包含EPON数据文件的文件夹")
        if folder_path: 
            self.input_path_var.set(folder_path)

    def select_output_dir(self):
        dir_path = filedialog.askdirectory(title="选择输出目录")
        if dir_path: 
            self.output_dir_var.set(dir_path)

    def select_local_output_dir(self):
        dir_path = filedialog.askdirectory(title="选择输出目录")
        if dir_path: 
            self.local_output_dir_var.set(dir_path)

    def clear_log(self):
        self.log_text.delete(1.0, tk.END)

    def log(self, message):
        self.log_queue.put(message)

    def update_log(self):
        try:
            while True:
                message = self.log_queue.get_nowait()
                self.log_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
                self.log_text.see(tk.END)
        except queue.Empty: 
            pass
        self.root.after(100, self.update_log)

    def parse_device_list(self, text):
        """解析设备列表文本，格式: 设备名-...-IP（IP在最后一组）"""
        devices = []
        # 匹配IPv4地址的正则
        ip_pattern = r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'
        
        for line in text.strip().split('\n'):
            line = line.strip()
            if not line:
                continue
            
            # 查找行中的IP地址
            ip_match = re.search(ip_pattern, line)
            if ip_match:
                ip = ip_match.group(1)
                # 设备名是IP之前的所有内容（去掉末尾的横杠）
                device_name = line[:ip_match.start()].rstrip('-').strip()
                if device_name:
                    devices.append((device_name, ip))
                else:
                    self.log(f"警告: 设备名为空: {line}")
            else:
                self.log(f"警告: 无法解析行（未找到IP）: {line}")
        
        return devices

    def start_ssh_collection(self):
        """开始SSH采集"""
        if not PARAMIKO_AVAILABLE:
            messagebox.showerror("错误", "未安装paramiko库，无法使用SSH功能。\n请运行: pip install paramiko")
            return
        
        # 获取参数
        device_text = self.device_text.get(1.0, tk.END).strip()
        username = self.username_var.get().strip()
        password = self.password_var.get().strip()
        output_dir = self.output_dir_var.get().strip()
        
        try:
            port = int(self.port_var.get().strip())
        except ValueError:
            messagebox.showerror("错误", "端口号必须是数字！")
            return
        
        # 验证参数
        if not device_text:
            messagebox.showerror("错误", "请输入设备列表！")
            return
        
        devices = self.parse_device_list(device_text)
        if not devices:
            messagebox.showerror("错误", "未能解析到任何设备，请检查格式！\n格式示例: 山下湖7606-[192.168.1.1]")
            return
        
        if not username or not password:
            messagebox.showerror("错误", "请输入SSH用户名和密码！")
            return
        
        if not output_dir:
            # 默认使用当前目录
            output_dir = os.getcwd()
            self.output_dir_var.set(output_dir)
        
        if not os.path.exists(output_dir):
            messagebox.showerror("错误", "输出目录不存在！")
            return
        
        # 重置停止标志，禁用开始按钮，启用停止按钮
        self.stop_flag = False
        self.ssh_btn.config(state=tk.DISABLED, text="⏳ 正在采集...")
        self.stop_ssh_btn.config(state=tk.NORMAL)
        self.status_var.set(f"正在采集 {len(devices)} 台设备...")
        
        self.ssh_thread = threading.Thread(
            target=self.ssh_collection_task, 
            args=(devices, username, password, port, output_dir), 
            daemon=True
        )
        self.ssh_thread.start()

    def stop_ssh_collection(self):
        """停止SSH采集"""
        self.stop_flag = True
        self.log("正在停止采集...")
        self.stop_ssh_btn.config(state=tk.DISABLED)
        self.status_var.set("正在停止...")

    def ssh_collection_task(self, devices, username, password, port, output_dir):
        """SSH采集任务（在后台线程中运行）"""
        try:
            total_devices = len(devices)
            processed_count = 0
            
            for idx, (device_name, ip) in enumerate(devices, 1):
                # 检查是否停止
                if self.stop_flag:
                    self.log("采集已停止")
                    break
                
                self.log(f"[{idx}/{total_devices}] 正在连接设备: {device_name} [{ip}]")
                
                try:
                    ssh = SSHConnection(ip, port, username, password)
                    ssh.connect(timeout=10)
                    self.log(f"  成功连接到 {device_name}")
                    
                    output_buffer = []
                    output_buffer.append(f"# 设备: {device_name}\n")
                    output_buffer.append(f"# IP: {ip}\n")
                    output_buffer.append(f"# 采集时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    output_buffer.append("="*60 + "\n")
                    
                    # 执行slot 2-7命令
                    for slot in range(2, 8):
                        # 检查是否停止
                        if self.stop_flag:
                            self.log("采集已停止")
                            break
                        
                        cmd = f"dis onu slot {slot}"
                        self.log(f"  执行命令: {cmd}")
                        
                        output = ssh.get_full_output(cmd, timeout=30)
                        
                        # 检查是否无板卡
                        if "Wrong parameter" in output:
                            self.log(f"    槽位 {slot} 无板卡，跳过")
                            continue
                        
                        # 检查是否有ONU信息
                        if "ONUs found:" in output:
                            match = re.search(r'ONUs found:\s*(\d+)', output)
                            if match:
                                onu_count = match.group(1)
                                self.log(f"    槽位 {slot} 发现 {onu_count} 个ONU")
                        
                        output_buffer.append(f"\n# 命令: {cmd}\n")
                        output_buffer.append(output)
                        output_buffer.append("\n" + "-"*60 + "\n")
                    
                    # 保存到文件
                    save_path = os.path.join(output_dir, f"{device_name}.txt")
                    with open(save_path, 'w', encoding='utf-8') as f:
                        f.write(''.join(output_buffer))
                    
                    self.log(f"  已保存: {save_path}")
                    ssh.close()
                    processed_count += 1
                    
                except Exception as e:
                    self.log(f"  设备 {device_name} 连接失败: {str(e)}")
                    continue
                
                # 检查是否停止
                if self.stop_flag:
                    break
            
            self.log("="*60)
            if self.stop_flag:
                self.log(f"SSH采集已停止！共处理 {processed_count}/{total_devices} 台设备")
            else:
                self.log(f"SSH采集完成！共处理 {processed_count} 台设备")
            self.log(f"输出目录: {output_dir}")
            
            if not self.stop_flag:
                self.root.after(0, lambda: messagebox.showinfo("完成", f"SSH采集完成！\n共处理 {processed_count} 台设备\n输出目录: {output_dir}"))
                self.root.after(0, lambda: os.startfile(output_dir))
            
        except Exception as e:
            self.log(f"严重错误: {traceback.format_exc()}")
            self.root.after(0, lambda: messagebox.showerror("错误", f"采集过程中发生错误:\n{str(e)}"))
        finally:
            self.root.after(0, lambda: self.ssh_btn.config(state=tk.NORMAL, text="🚀 开始SSH采集"))
            self.root.after(0, lambda: self.stop_ssh_btn.config(state=tk.DISABLED))
            self.root.after(0, lambda: self.status_var.set("就绪"))

    def start_processing(self):
        """开始本地文件处理"""
        input_path = self.input_path_var.get().strip()
        output_dir = self.local_output_dir_var.get().strip()
        
        if not input_path or not os.path.exists(input_path):
            messagebox.showerror("错误", "输入路径无效！")
            return
        if not output_dir or not os.path.exists(output_dir):
            messagebox.showerror("错误", "输出目录无效！")
            return
        
        self.process_btn.config(state=tk.DISABLED, text="⏳ 正在统计数据...")
        self.status_var.set("正在处理...")
        threading.Thread(target=self.process_task, args=(input_path, output_dir), daemon=True).start()

    def process_task(self, input_path, output_dir):
        """本地文件处理任务"""
        try:
            files = [input_path] if os.path.isfile(input_path) else [os.path.join(input_path, f) for f in os.listdir(input_path) if f.lower().endswith('.txt')]
            if not files: 
                raise ValueError("未找到TXT文件")

            for file_path in files:
                self.log(f"解析文件: {os.path.basename(file_path)}")
                slot_data = self.parse_epon_data(file_path)
                output_filename = os.path.splitext(os.path.basename(file_path))[0] + ".xlsx"
                self.generate_excel_report(slot_data, os.path.join(output_dir, output_filename))
            
            self.log(f"处理完成！")
            self.root.after(0, lambda: messagebox.showinfo("完成", "报表生成完毕。"))
            self.root.after(0, lambda: os.startfile(output_dir))
        except Exception as e:
            self.log(f"严重错误: {traceback.format_exc()}")
        finally:
            self.root.after(0, lambda: self.process_btn.config(state=tk.NORMAL, text="🚀 开始处理并生成Excel"))
            self.root.after(0, lambda: self.status_var.set("就绪"))

    def parse_epon_data(self, file_path):
        """解析EPON数据文件"""
        slot_data = {s: {p: {'在线': 0, '离线': 0, '静默': 0} for p in range(1, 25)} for s in range(2, 8)}
        current_slot, current_pon = None, None
        content = None
        for enc in ['utf-8', 'gbk', 'gb2312']:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    content = f.readlines()
                break
            except: 
                continue
        if not content: 
            raise ValueError("无法读取文件")

        for line in content:
            line = line.strip()
            if 'dis onu slot' in line:
                match = re.search(r'dis onu slot\s+(\d+)', line)
                if match: 
                    current_slot = int(match.group(1))
                continue
            if current_slot and 2 <= current_slot <= 7 and 'Olt' in line and '/0/' in line:
                match = re.search(r'Olt\d+/0/(\d+)', line)
                if match: 
                    current_pon = int(match.group(1))
                continue
            if current_slot and current_pon and line and not line.startswith('-'):
                if any(k in line for k in ['State', 'MAC', 'LOID', 'LLID', 'Port']): 
                    continue
                parts = re.split(r'\s+', line)
                if len(parts) >= 2:
                    state = parts[-2]
                    key = '在线' if state == 'Up' else '离线' if state == 'Offline' else '静默' if state == 'Silent' else None
                    if key: 
                        slot_data[current_slot][current_pon][key] += 1
        return slot_data

    def generate_excel_report(self, slot_data, output_path):
        """生成Excel报表"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EPON统计报表"

        # 样式定义
        color_slot_name = PatternFill(start_color="FDE9D9", fill_type="solid")
        color_pon_header = PatternFill(start_color="D9E1F2", fill_type="solid")
        color_idle_yes = PatternFill(start_color="FFFF00", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal="center", vertical="center")

        total_idle_count = 0

        # 1. 标题
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f"统计信息(生成日期: {datetime.now().strftime('%Y-%m-%d')})"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = center_align

        # 2. 列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        for col in range(3, 15):
            ws.column_dimensions[get_column_letter(col)].width = 5

        current_row = 2
        slot_names = {2: "2号槽位", 3: "3号槽位", 4: "4号槽位", 5: "5号槽位", 6: "6号槽位", 7: "7号槽位"}

        # 3. 槽位数据循环 - 只处理txt中有数据的槽位
        for slot_num in range(2, 8):
            slot_info = slot_data.get(slot_num, {})
            
            # 检查该槽位是否有数据（是否有任何PON口有非零值）
            has_data = False
            for pon_id in range(1, 25):
                pon_data = slot_info.get(pon_id, {})
                if pon_data.get('在线', 0) > 0 or pon_data.get('离线', 0) > 0 or pon_data.get('静默', 0) > 0:
                    has_data = True
                    break
            
            # 如果该槽位没有数据，跳过不登记
            if not has_data:
                continue
            
            start_merge_row = current_row
            
            for group in [range(1, 24, 2), range(2, 25, 2)]:
                rows = [("PON", None), ("在线", "在线"), ("离线", "离线"), ("静默", "静默"), ("空闲", "空闲")]
                for label, data_key in rows:
                    ws.cell(row=current_row, column=2, value=label).alignment = center_align
                    for idx, pon_id in enumerate(group, start=3):
                        cell = ws.cell(row=current_row, column=idx)
                        cell.alignment = center_align
                        if label == "PON":
                            cell.value = pon_id
                            cell.fill = color_pon_header
                            ws.cell(row=current_row, column=2).fill = color_pon_header
                        elif label == "空闲":
                            is_idle = slot_info.get(pon_id, {}).get('在线', 0) == 0
                            if is_idle:
                                cell.value = "是"
                                cell.fill = color_idle_yes
                                cell.font = Font(bold=True)
                                total_idle_count += 1
                            else:
                                cell.value = "否"
                        else:
                            cell.value = slot_info.get(pon_id, {}).get(data_key, 0)
                    current_row += 1
            
            ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=current_row-1, end_column=1)
            slot_cell = ws.cell(row=start_merge_row, column=1, value=slot_names[slot_num])
            slot_cell.alignment = center_align
            slot_cell.fill = color_slot_name
            slot_cell.font = Font(bold=True)

        # 4. 给表格主体添加边框
        for r in range(1, current_row):
            for c in range(1, 15):
                ws.cell(row=r, column=c).border = thin_border

        # 5. 统计行
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        stat_cell = ws.cell(row=current_row, column=1)
        stat_cell.value = f"截止{datetime.now().strftime('%Y年%m月%d日')}统计该设备可利用PON口数量：{total_idle_count}"
        stat_cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")
        stat_cell.alignment = center_align
        stat_cell.font = Font(bold=True)

        # 6. 备注
        notes = [
            "", "备注：",
            "1. 空闲一栏标记为「是」，说明该口下无在线用户。需留意离线和静默数量。",
            "2. 离线：若确认为撤销点位请反馈技术部删除配置；FTTH日常关机则无需处理。",
            "3. 静默：说明有ONU在线但未配置业务，请及时核实并下发配置。",
            "4. 统计结果以发布日期当天为准。"
        ]
        for note in notes:
            current_row += 1
            ws.cell(row=current_row, column=1, value=note).font = Font(size=10)

        wb.save(output_path)


if __name__ == "__main__":
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except: 
        pass
    
    root = tk.Tk()
    app = EPONPortAnalyzer(root)
    root.mainloop()
